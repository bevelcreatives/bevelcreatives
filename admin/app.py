"""
B's Bobux - Admin Dashboard
A single Flask app that reads ticket_data.json (shared with the Discord bot)
and serves an admin-only UI for orders, analytics, and XLSX exports.

Env vars:
  ADMIN_USERNAME    - required login username
  ADMIN_PASSWORD    - required login password (plaintext, kept in env only)
  SECRET_KEY        - Flask session secret (random string)
  TICKET_DATA_PATH  - path to the bot's ticket_data.json  (default: ./ticket_data.json)
  HOST, PORT        - where Flask listens (default 0.0.0.0:8000)
"""
from __future__ import annotations

import base64
import hmac
import io
import json
import os
from datetime import datetime
from functools import wraps
from pathlib import Path
from urllib.parse import urlparse

import requests as _requests

# Load .env before reading env vars (python-dotenv is in requirements.txt)
try:
    from dotenv import load_dotenv
    APP_DIR = Path(__file__).resolve().parent
    load_dotenv(APP_DIR / ".env")
    load_dotenv(APP_DIR.parent / ".env", override=True)
except ImportError:
    pass

from flask import (
    Flask, Response, abort, jsonify, redirect, render_template,
    request, send_file, session, url_for, flash,
)

from .data import (
    all_orders, filter_orders, analytics, user_history, leaderboard,
    button_stats, ineligible_choice_stats, clear_dashboard_data, STATUSES, STATUS_LABELS,
)

# â"€â"€â"€â"€â"€â"€â"€â"€â"€ openpyxl (optional import; raise a helpful error if missing) â"€â"€â"€â"€â"€â"€â"€â"€â"€
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# â"€â"€â"€â"€â"€â"€â"€â"€â"€ app â"€â"€â"€â"€â"€â"€â"€â"€â"€
app = Flask(__name__, static_folder="static", template_folder="templates")
DEFAULT_SECRET_KEY = "dev-insecure-change-me-in-production"
IS_PRODUCTION = bool(os.getenv("RENDER") or os.getenv("VERCEL") or os.getenv("FLASK_ENV") == "production")

app.secret_key = os.getenv("SECRET_KEY", DEFAULT_SECRET_KEY)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
if IS_PRODUCTION:
    app.config["SESSION_COOKIE_SECURE"] = True
# When deployed behind a reverse-proxy prefix (e.g. /admin via Vercel rewrite)
if os.getenv("SESSION_COOKIE_DOMAIN"):
    app.config["SESSION_COOKIE_DOMAIN"] = os.getenv("SESSION_COOKIE_DOMAIN")
if os.getenv("SESSION_COOKIE_PATH"):
    app.config["SESSION_COOKIE_PATH"] = os.getenv("SESSION_COOKIE_PATH")

ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")
_KV_URL        = os.getenv("KV_REST_API_URL", "")
_KV_TOKEN      = os.getenv("KV_REST_API_TOKEN", "")
_BOT_API_URL   = os.getenv("BOT_API_URL", "")
_BOT_API_SECRET = os.getenv("BOT_API_SECRET", "")

if IS_PRODUCTION and app.secret_key == DEFAULT_SECRET_KEY:
    raise RuntimeError("SECRET_KEY must be set in production.")

if IS_PRODUCTION and not ADMIN_PASSWORD:
    raise RuntimeError("ADMIN_PASSWORD must be set in production.")


def _is_safe_redirect_target(target: str | None) -> bool:
    if not target:
        return False
    parsed = urlparse(target)
    return not parsed.scheme and not parsed.netloc and target.startswith("/") and not target.startswith("//")


def _safe_redirect_target() -> str:
    target = request.args.get("next", "")
    return target if _is_safe_redirect_target(target) else url_for("dashboard")


# â"€â"€â"€â"€â"€â"€â"€â"€â"€ auth â"€â"€â"€â"€â"€â"€â"€â"€â"€
def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("authed"):
            return redirect(url_for("login", next=request.path))
        return fn(*args, **kwargs)
    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "")
        if not ADMIN_PASSWORD:
            flash("Server is misconfigured: ADMIN_PASSWORD is not set.", "error")
        elif hmac.compare_digest(u, ADMIN_USERNAME) and hmac.compare_digest(p, ADMIN_PASSWORD):
            session["authed"] = True
            session["user"] = u
            return redirect(_safe_redirect_target())
        else:
            flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# â"€â"€â"€â"€â"€â"€â"€â"€â"€ views â"€â"€â"€â"€â"€â"€â"€â"€â"€
@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html",
                           statuses=STATUSES, status_labels=STATUS_LABELS)


@app.route("/analytics")
@login_required
def analytics_page():
    return render_template("analytics.html")


@app.route("/users")
@login_required
def users_page():
    return render_template("users.html")


@app.route("/leaderboard")
@login_required
def leaderboard_page():
    return render_template("leaderboard.html")


# â"€â"€â"€â"€â"€â"€â"€â"€â"€ JSON API â"€â"€â"€â"€â"€â"€â"€â"€â"€
@app.route("/api/orders")
@login_required
def api_orders():
    q         = request.args.get("q", "")
    status    = request.args.get("status", "")
    date_from = request.args.get("from", "")
    date_to   = request.args.get("to", "")
    orders = filter_orders(all_orders(), q=q, status=status,
                           date_from=date_from, date_to=date_to)
    # Trim heavy fields for list view
    light = []
    for o in orders:
        light.append({
            "order":              o["order"],
            "created_at_display": o["created_at_display"],
            "created_at":         o["created_at"],
            "roblox":             o["roblox"],
            "roblox_display_name": o.get("roblox_display_name", ""),
            "roblox_edited":      o["roblox_edited"],
            "amount":             o["amount"],
            "amount_edited":      o["amount_edited"],
            "discord_name":       o["discord_name"],
            "discord_user_id":    o["discord_user_id"],
            "status":             o["status"],
            "status_label":       STATUS_LABELS.get(o["status"], o["status"]),
        })
    return jsonify({"orders": light, "count": len(light),
                    "statuses": STATUSES, "status_labels": STATUS_LABELS})


@app.route("/api/order/<int:order_no>")
@login_required
def api_order_detail(order_no: int):
    match = next((o for o in all_orders() if o["order"] == order_no), None)
    if not match:
        return jsonify({"error": "not_found"}), 404
    match = dict(match)
    match["status_label"] = STATUS_LABELS.get(match["status"], match["status"])
    return jsonify(match)


@app.route("/api/order/<int:order_no>/screenshot")
@login_required
def api_order_screenshot(order_no: int):
    match = next((o for o in all_orders() if o["order"] == order_no), None)
    if not match:
        return jsonify({"error": "not_found"}), 404

    # Try KV first (uploaded by bot on screenshot receipt)
    if _KV_URL and _KV_TOKEN:
        try:
            resp = _requests.get(
                f"{_KV_URL}/get/screenshot:{order_no}",
                headers={"Authorization": f"Bearer {_KV_TOKEN}"},
                timeout=5,
            )
            if resp.ok:
                value = resp.json().get("result")
                if value:
                    d = json.loads(value)
                    img_bytes = base64.b64decode(d["data"])
                    content_type = d.get("content_type", "image/png")
                    return send_file(
                        io.BytesIO(img_bytes),
                        mimetype=content_type,
                        as_attachment=False,
                    )
        except Exception:
            pass

    # Fall back to local file (local dev only)
    filename = match.get("screenshot_filename")
    if filename:
        screenshot_dir = Path(__file__).resolve().parent.parent / "screenshots"
        file_path = screenshot_dir / filename
        if file_path.exists():
            return send_file(file_path)

    # Last resort: CDN URL (may be expired)
    target = match.get("screenshot_log_url") or match.get("screenshot_url")
    if not target:
        return jsonify({"error": "screenshot_not_found"}), 404

    parsed = urlparse(target)
    allowed_hosts = {
        "cdn.discordapp.com",
        "media.discordapp.net",
        "images-ext-1.discordapp.net",
        "images-ext-2.discordapp.net",
    }
    if parsed.scheme != "https" or parsed.netloc not in allowed_hosts:
        abort(400, description="Invalid screenshot URL.")

    return redirect(target, code=302)


@app.route("/api/analytics")
@login_required
def api_analytics():
    return jsonify(analytics(
        date_from=request.args.get("from", ""),
        date_to=request.args.get("to", ""),
    ))


@app.route("/api/user")
@login_required
def api_user():
    result = user_history(
        q=request.args.get("q", ""),
        date_from=request.args.get("from", ""),
        date_to=request.args.get("to", ""),
    )
    for m in result.get("matches", []):
        m["status_label"] = STATUS_LABELS.get(m.get("status"), m.get("status"))
    return jsonify(result)


@app.route("/api/leaderboard")
@login_required
def api_leaderboard():
    return jsonify(leaderboard(
        date_from=request.args.get("from", ""),
        date_to=request.args.get("to", ""),
    ))


@app.route("/api/button-stats")
@login_required
def api_button_stats():
    return jsonify(button_stats(
        date_from=request.args.get("from", ""),
        date_to=request.args.get("to", ""),
    ))


@app.route("/api/ineligible-choices")
@login_required
def api_ineligible_choices():
    return jsonify(ineligible_choice_stats(
        date_from=request.args.get("from", ""),
        date_to=request.args.get("to", ""),
    ))


@app.route("/api/clear-data", methods=["POST"])
@login_required
def api_clear_data():
    payload = request.get_json(silent=True) or {}
    if str(payload.get("key", "")).strip() != "5040":
        return jsonify({"error": "invalid_key"}), 403

    clear_dashboard_data()
    return jsonify({"ok": True})


@app.route("/api/bulk-complete", methods=["POST"])
@login_required
def api_bulk_complete():
    if not _BOT_API_URL:
        return jsonify({"error": "bot_api_not_configured",
                        "message": "Set BOT_API_URL in the dashboard environment."}), 503
    payload = request.get_json(silent=True) or {}
    orders  = payload.get("orders", [])
    if not orders:
        return jsonify({"error": "no_orders"}), 400
    headers = {"Content-Type": "application/json"}
    if _BOT_API_SECRET:
        headers["Authorization"] = f"Bearer {_BOT_API_SECRET}"
    try:
        resp = _requests.post(
            _BOT_API_URL.rstrip("/") + "/api/bulk-complete",
            json={"orders": orders, "completed_by": "Dashboard Admin"},
            headers=headers,
            timeout=60,
        )
        return jsonify(resp.json()), resp.status_code
    except Exception as exc:
        return jsonify({"error": "bot_unreachable", "message": str(exc)}), 502


# â"€â"€â"€â"€â"€â"€â"€â"€â"€ XLSX export â"€â"€â"€â"€â"€â"€â"€â"€â"€
def _require_openpyxl():
    if not HAS_OPENPYXL:
        abort(500, description="openpyxl is not installed on the server.")


def _autosize(ws):
    for col in ws.columns:
        max_len = 8
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(60, len(str(cell.value))))
        ws.column_dimensions[letter].width = max_len + 2


def _write_headers(ws, headers: list[str]):
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _xlsx_response(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/orders.xlsx")
@login_required
def export_orders():
    _require_openpyxl()
    q         = request.args.get("q", "")
    status    = request.args.get("status", "")
    date_from = request.args.get("from", "")
    date_to   = request.args.get("to", "")
    orders = filter_orders(all_orders(), q=q, status=status,
                           date_from=date_from, date_to=date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = ["Order #", "Status", "Created", "Roblox", "Amount (Robux)",
               "Discord Username", "Discord ID", "Pre-Order", "Eligible On",
               "Completed At", "Cancelled At", "Rejected At", "Auto-Deleted At",
               "Screenshot Received At", "Screenshot URL"]
    _write_headers(ws, headers)
    for i, o in enumerate(orders, start=2):
        ws.cell(row=i, column=1,  value=o["order"])
        ws.cell(row=i, column=2,  value=STATUS_LABELS.get(o["status"], o["status"]))
        ws.cell(row=i, column=3,  value=o["created_at_display"])
        ws.cell(row=i, column=4,  value=o["roblox"])
        ws.cell(row=i, column=5,  value=o["amount"])
        ws.cell(row=i, column=6,  value=o["discord_name"])
        ws.cell(row=i, column=7,  value=o["discord_user_id"])
        ws.cell(row=i, column=8,  value="Yes" if o["is_preorder"] else "No")
        ws.cell(row=i, column=9,  value=o["eligible_on"])
        ws.cell(row=i, column=10, value=o["completed_at_display"])
        ws.cell(row=i, column=11, value=o["cancelled_at_display"])
        ws.cell(row=i, column=12, value=o["rejected_at_display"])
        ws.cell(row=i, column=13, value=o["auto_deleted_at_display"])
        ws.cell(row=i, column=14, value=o["screenshot_at_display"])
        ws.cell(row=i, column=15, value=o["screenshot_url"])
    _autosize(ws)

    # Second sheet: audit logs
    ws2 = wb.create_sheet("Audit Log")
    _write_headers(ws2, ["Order #", "Roblox", "Status", "Event"])
    row = 2
    for o in orders:
        events = o.get("log_history") or []
        if not events:
            ws2.cell(row=row, column=1, value=o["order"])
            ws2.cell(row=row, column=2, value=o["roblox"])
            ws2.cell(row=row, column=3, value=STATUS_LABELS.get(o["status"], o["status"]))
            ws2.cell(row=row, column=4, value="(no events)")
            row += 1
        else:
            for ev in events:
                ws2.cell(row=row, column=1, value=o["order"])
                ws2.cell(row=row, column=2, value=o["roblox"])
                ws2.cell(row=row, column=3, value=STATUS_LABELS.get(o["status"], o["status"]))
                ws2.cell(row=row, column=4, value=ev)
                row += 1
    _autosize(ws2)

    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return _xlsx_response(wb, f"orders-{stamp}.xlsx")


@app.route("/export/user.xlsx")
@login_required
def export_user():
    _require_openpyxl()
    result = user_history(
        q=request.args.get("q", ""),
        date_from=request.args.get("from", ""),
        date_to=request.args.get("to", ""),
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _write_headers(ws, ["Query", "Total Orders", "Completed Orders",
                        "Cancelled Orders", "Total Amount (Completed, Robux)"])
    ws.append([
        result["query"], result["totals"].get("orders", 0),
        result["totals"].get("completed", 0),
        result["totals"].get("cancelled", 0),
        result["totals"].get("amount_total_completed", 0),
    ])
    _autosize(ws)

    ws2 = wb.create_sheet("Orders")
    _write_headers(ws2, ["Order #", "Status", "Created", "Roblox", "Amount (Robux)",
                         "Discord", "Discord ID", "Completed At"])
    for i, o in enumerate(result["matches"], start=2):
        ws2.cell(row=i, column=1, value=o["order"])
        ws2.cell(row=i, column=2, value=STATUS_LABELS.get(o["status"], o["status"]))
        ws2.cell(row=i, column=3, value=o["created_at_display"])
        ws2.cell(row=i, column=4, value=o["roblox"])
        ws2.cell(row=i, column=5, value=o["amount"])
        ws2.cell(row=i, column=6, value=o["discord_name"])
        ws2.cell(row=i, column=7, value=o["discord_user_id"])
        ws2.cell(row=i, column=8, value=o["completed_at_display"])
    _autosize(ws2)

    safe_q = (result["query"] or "user").replace(" ", "_")
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return _xlsx_response(wb, f"user-{safe_q}-{stamp}.xlsx")


@app.route("/export/analytics.xlsx")
@login_required
def export_analytics():
    _require_openpyxl()
    date_from = request.args.get("from", "")
    date_to   = request.args.get("to", "")
    stats  = analytics(date_from=date_from, date_to=date_to)
    btn    = button_stats(date_from=date_from, date_to=date_to)
    ic     = ineligible_choice_stats(date_from=date_from, date_to=date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _write_headers(ws, ["Metric", "Value"])
    for r in [
        ("From",                              stats["from"]),
        ("To",                                stats["to"]),
        ("Total Tickets Opened",              stats["total_orders"]),
        ("Completed",                         stats["total_completed"]),
        ("Cancelled",                         stats["total_cancelled"]),
        ("Rejected",                          stats["total_rejected"]),
        ("Auto-deleted",                      stats["total_auto_deleted"]),
        ("Still Open",                        stats["total_open"]),
        ("Awaiting Review",                   stats["total_awaiting"]),
        ("Total Amount (Robux, completed)",   stats["total_amount_completed"]),
        ("",                                  ""),
        ("Button Clicks - Total",             btn["totals"].get("total", 0)),
        ("Button Clicks - Not In Group",      btn["totals"].get("not_joined", 0)),
        ("Button Clicks - Not Eligible Yet",  btn["totals"].get("ineligible", 0)),
        ("Button Clicks - Eligible",          btn["totals"].get("eligible", 0)),
        ("Button Clicks - Not Found",         btn["totals"].get("not_found", 0)),
        ("",                                  ""),
        ("Ineligible - Proceeded",            ic["totals"].get("proceed", 0)),
        ("Ineligible - Bought Later",         ic["totals"].get("later", 0)),
    ]:
        ws.append(list(r))
    _autosize(ws)

    ws2 = wb.create_sheet("Daily")
    _write_headers(ws2, ["Date", "Opened", "Completed", "Cancelled",
                         "Rejected", "Auto-deleted", "Amount (Completed Robux)"])
    for i, d in enumerate(stats["days"], start=2):
        ws2.cell(row=i, column=1, value=d)
        ws2.cell(row=i, column=2, value=stats["series"]["opened"][i-2])
        ws2.cell(row=i, column=3, value=stats["series"]["completed"][i-2])
        ws2.cell(row=i, column=4, value=stats["series"]["cancelled"][i-2])
        ws2.cell(row=i, column=5, value=stats["series"]["rejected"][i-2])
        ws2.cell(row=i, column=6, value=stats["series"]["auto_deleted"][i-2])
        ws2.cell(row=i, column=7, value=stats["series"]["amount"][i-2])
    _autosize(ws2)

    ws3 = wb.create_sheet("Button Clicks")
    _write_headers(ws3, ["Date", "Total Clicks", "Not In Group",
                         "Not Eligible Yet", "Eligible", "Not Found"])
    for i, d in enumerate(btn["days"], start=2):
        s = btn["series"]
        ws3.cell(row=i, column=1, value=d)
        ws3.cell(row=i, column=2, value=s["total"][i-2])
        ws3.cell(row=i, column=3, value=s["not_joined"][i-2])
        ws3.cell(row=i, column=4, value=s["ineligible"][i-2])
        ws3.cell(row=i, column=5, value=s["eligible"][i-2])
        ws3.cell(row=i, column=6, value=s["not_found"][i-2])
    _autosize(ws3)

    ws4 = wb.create_sheet("Ineligible Choices")
    _write_headers(ws4, ["Date", "Proceeded", "Bought Later"])
    for i, d in enumerate(ic["days"], start=2):
        ws4.cell(row=i, column=1, value=d)
        ws4.cell(row=i, column=2, value=ic["series"]["proceed"][i-2])
        ws4.cell(row=i, column=3, value=ic["series"]["later"][i-2])
    _autosize(ws4)

    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return _xlsx_response(wb, f"analytics-{stamp}.xlsx")


# â"€â"€â"€â"€â"€â"€â"€â"€â"€ health â"€â"€â"€â"€â"€â"€â"€â"€â"€
@app.route("/healthz")
def healthz():
    return "ok", 200


@app.after_request
def add_security_headers(resp: Response):
    resp.headers.setdefault("X-Frame-Options", "DENY")
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("Referrer-Policy", "same-origin")
    resp.headers.setdefault("Cache-Control", "no-store")
    resp.headers["Content-Security-Policy"] = (
        "default-src 'self'; img-src 'self' https: data:; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "connect-src 'self'; frame-ancestors 'none'; base-uri 'self'; form-action 'self'"
    )
    return resp


# â"€â"€ WSGI entrypoint â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
# When SCRIPT_NAME is set (e.g. "/admin"), mount the app at that prefix so
# Flask's url_for() and redirect() include it automatically.
_script_name = os.getenv("SCRIPT_NAME", "")
if _script_name:
    from werkzeug.middleware.dispatcher import DispatcherMiddleware
    from werkzeug.exceptions import NotFound as _NotFound
    wsgi_app = DispatcherMiddleware(_NotFound(), {_script_name: app})
else:
    wsgi_app = app

if __name__ == "__main__":
    host = os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "8000"))
    app.run(host=host, port=port, debug=os.getenv("FLASK_DEBUG") == "1")

